"""
读取登录密码 Excel 文件 - 完整信息
"""
import openpyxl
import os

excel_path = r"C:\Users\15606\Desktop\AI_Project\纳税申报\纳税申报模块\员工登录方式\登录密码.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print("Excel 内容：")
print("=" * 60)

# 读取表头
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    headers.append(cell_value if cell_value else "")
    print(f"列{col}: {cell_value}")

print("=" * 60)

# 读取数据行
for row in range(2, ws.max_row + 1):
    print(f"\n第{row}行数据:")
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row, column=col).value
        header = headers[col-1] if col-1 < len(headers) else f"列{col}"
        print(f"  {header}: {cell_value}")
