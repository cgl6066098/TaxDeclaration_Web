"""
浙江省电子税务局 - 登录模块
从 Excel 读取登录信息，自动完成登录流程

功能集成：
1. 检测现有浏览器 - 使用 CDP 端口检测 Edge 是否运行
2. 连接现有浏览器 - 通过 CDP 连接到已有 Edge
3. 多标签管理 - 列出所有页面，找到匹配的标签
4. 智能打开 - 不在税务局页面时打开新标签

永久保持：
5. 永久保持浏览器 - 脚本退出后浏览器继续运行
6. 不设置 timeout - 没有任何超时限制

使用方法:
    python 登录.py                    # 默认：前台保持打开
    python 登录.py --headless         # 后台无头模式
    python 登录.py --no-keep-alive    # 不保持打开
"""
import sys
import os

# 添加模块路径
sys.path.insert(0, os.path.dirname(__file__))

from browser import (
    get_page, 
    get_page_by_url, 
    screenshot, 
    wait, 
    set_keep_alive, 
    list_pages,
    browser_manager
)
import time
import openpyxl
import argparse


# 浙江省电子税务局 URL
TAX_URL = "https://tpass.zhejiang.chinatax.gov.cn:8443/#/login?redirect_uri=https%3A%2F%2Fetax.zhejiang.chinatax.gov.cn%3A8443%2Fmhzx%2Fapi%2Fmh%2Ftpass%2Fcode&client_id=tct8zta97w6c46zdt9zc2648227df5z2&response_type=code&state=232c3847df0f4f36a00c1b5e55ca3fe9"


class TaxLogin:
    """税务登录类"""

    def __init__(self):
        self.page = None

        # 登录信息
        self.shehui_xinyong = ""
        self.phone = ""
        self.password = ""

        # 目录配置
        self.base_dir = os.path.dirname(__file__)
        self.excel_path = os.path.join(
            self.base_dir,
            "员工登录方式",
            "登录密码.xlsx"
        )

    def read_login_info(self) -> 'TaxLogin':
        """从 Excel 读取登录信息"""
        print("\n" + "=" * 60)
        print("[1/4] 读取登录信息")
        print("=" * 60)

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value or "")

        for col in range(1, ws.max_column + 1):
            header = headers[col-1]
            value = ws.cell(row=2, column=col).value

            if "纳税人识别号" in header or "社会信用代码" in header:
                self.shehui_xinyong = str(value)
            elif "手机" in header:
                self.phone = str(value)
            elif "密码" in header:
                self.password = str(value)

        print(f"社会信用代码：{self.shehui_xinyong}")
        print(f"手机号码：{self.phone}")
        print(f"密码：{'*' * len(self.password)}")

        return self

    def start(self, headless: bool = False, keep_alive: bool = True) -> 'TaxLogin':
        """
        获取或创建浏览器
        
        功能：
        1. 检测现有浏览器 - 使用 CDP 端口检测 Edge 是否运行
        2. 连接现有浏览器 - 通过 CDP 连接到已有 Edge
        3. 多标签管理 - 列出所有页面
        4. 智能打开 - 不在税务局页面时打开新标签
        """
        print("\n" + "=" * 60)
        print("[2/4] 获取浏览器")
        print("=" * 60)

        set_keep_alive(keep_alive)
        
        # 获取浏览器（自动检测并连接现有浏览器）
        self.page = get_page(headless=headless)
        
        # 显示当前浏览器状态
        print(f"\n当前 URL: {self.page.url}")
        print(f"当前标题：{self.page.title()}")
        
        # 检查是否在税务局页面
        if "zhejiang.chinatax.gov.cn" not in self.page.url:
            print("\n[提示] 当前不在税务局页面，打开/切换到税务局标签...")
            self.page = get_page_by_url("zhejiang.chinatax.gov.cn")
        
        # 列出所有页面
        pages = list_pages()
        print(f"\n当前浏览器标签页 ({len(pages)}个):")
        for i, p in enumerate(pages):
            is_current = "← 当前" if p["url"] == self.page.url else ""
            print(f"  [{i}] {p['title'][:30]}... {p['url'][:50]} {is_current}")
        
        print("\n[OK] 浏览器已就绪")
        return self

    def login(self) -> 'TaxLogin':
        """执行登录流程"""
        print("\n" + "=" * 60)
        print("[3/4] 执行登录")
        print("=" * 60)

        print(f"\n[1/6] 访问登录页面...")
        self.page.goto(TAX_URL)
        self.page.wait_for_load_state("networkidle")
        time.sleep(2)
        screenshot("01_login_page")
        print("        完成")

        print(f"[2/6] 点击 '代理业务'...")
        self.page.locator('text=代理业务').first.click()
        time.sleep(1)
        screenshot("02_daili")
        print("        完成")

        print(f"[3/6] 填写社会信用代码：{self.shehui_xinyong}")
        self.page.locator('input[placeholder*="代理机构统一社会信用代码"]').first.fill(self.shehui_xinyong)
        time.sleep(0.5)
        print("        完成")

        print(f"[4/6] 填写手机号码：{self.phone}")
        self.page.locator('input[placeholder*="手机号码"]').first.fill(self.phone)
        time.sleep(0.5)
        print("        完成")

        print(f"[5/6] 填写密码：{'*' * len(self.password)}")
        self.page.locator('input[type="password"]').first.fill(self.password)
        time.sleep(0.5)
        screenshot("03_filled")
        print("        完成")

        print(f"[6/6] 点击 '登录' 按钮...")
        self.page.locator('button:has-text("登录")').first.click()
        time.sleep(3)
        screenshot("04_after_login")
        print("        完成")

        print("\n[OK] 登录流程完成")
        return self

    def wait_for_verify(self, auto: bool = False):
        """等待验证码完成"""
        print("\n" + "=" * 60)
        print("[4/4] 验证码")
        print("=" * 60)
        
        if auto:
            print("\n自动验证模式 - 等待 5 秒...")
            time.sleep(5)
        else:
            print("验证码已显示，请手动完成")
            print("\n完成后按回车键继续...")
            input("\n按回车键继续...")

        screenshot("verified")
        time.sleep(3)
        screenshot("dashboard")
        
        # 显示登录后的页面
        print(f"\n当前 URL: {self.page.url}")
        print(f"页面标题：{self.page.title()}")
        
        return self

    def wait(self):
        """保持运行（浏览器不关闭）"""
        print("\n" + "=" * 70)
        print("浏览器保持打开 - 可以手动操作或调试")
        print("=" * 70)
        wait()


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='浙江省电子税务局登录',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python 登录.py                      # 前台保持打开（默认）
  python 登录.py --headless           # 后台无头模式
  python 登录.py --no-keep-alive      # 测试后关闭浏览器
  python 登录.py --auto-verify        # 自动完成验证码
        """
    )
    parser.add_argument('--headless', action='store_true', help='无头模式（后台运行，不显示窗口）')
    parser.add_argument('--keep-alive', action='store_true', default=True, help='保持浏览器打开（默认）')
    parser.add_argument('--no-keep-alive', action='store_false', dest='keep_alive', help='不保持浏览器打开（完成后关闭）')
    parser.add_argument('--auto-verify', action='store_true', help='自动完成验证码（跳过手动确认）')
    return parser.parse_args()


def 登录 (headless: bool = False, keep_alive: bool = True, auto_verify: bool = False):
    """
    快捷登录函数
    
    Args:
        headless: 是否无头模式
        keep_alive: 是否保持浏览器打开（默认 True）
        auto_verify: 是否自动完成验证码
    """
    login = TaxLogin()
    login.read_login_info()
    login.start(headless=headless, keep_alive=keep_alive)
    login.login()
    login.wait_for_verify(auto=auto_verify)

    if keep_alive:
        login.wait()


if __name__ == "__main__":
    args = parse_args()
    登录 (headless=args.headless, keep_alive=args.keep_alive, auto_verify=args.auto_verify)
